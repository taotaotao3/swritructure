# -*- coding: utf-8 -*-
import os , sys , codecs
import pdb
import re
import glob
#Program 1 | Library settings
import pathlib
import os
import re
import sys
import openpyxl as px
import pandas as pd
import pyperclip as clp
import csv
from openpyxl.styles import Alignment    # Alignmentクラスをインポート
from openpyxl.styles.borders import Border, Side # 罫線
from pathlib import Path

def get_structure(arg_1):

    print('sys.argv[1]:', str(arg_1))

    def from_file_list_export(start_file):
        # File extraction list with all directories below
        all_directory_list = []

        #Program 2 ｜ main procedure

        Folderpath = os.getcwd()#Specify the target folder


        #Program 3 | Functions for writing folders and files for each hierarchy
        def GetFolderFileNames(path, kaiso, all_directory_list_arg):
            # files = pathlib.Path(path).glob('*')#Get folders and files in that hierarchy
            files = os.listdir(path)#Get folders and files in that hierarchy
            all_directory_list_ch = []

            # Program 4 | Export folders and files one by one
            for file in files:
                output = file
                all_directory_list_ch.append(str(path) + '\\' + output)
                all_directory_list_ch.append(kaiso)
                all_directory_list_arg.append(all_directory_list_ch)
                all_directory_list_ch = []
                # In the case of Program 5 | Folder, call GetFolderFileNames (Program 6) for the folders and files in the hierarchy below it.
                #if not file[:3] == '.xml' or file[:3] == '.csv' or file[:3] == '.xml' or file[:3] == '.xml' or file[:3] == '.xml' or :
                if os.path.isdir(file) == True:

                    # print('file:',file)
                    GetFolderFileNames(path + "\\" + file, kaiso+1, all_directory_list_arg)
            return all_directory_list_arg

        all_directory_list = GetFolderFileNames(Folderpath, 0, all_directory_list)#GetFolderFileNamesCall (program 3)

        files = glob.glob("./*")

        for file in files:
            print(file)

        # Fill the list with the list of files in the current directory
        current_directory_files_list = os.listdir('./')


        ############################################################
        #!/usr/bin/python
        # -*- coding: utf-8 -*-
        # vim: fileencoding=utf-8

        grep_result = []
        def grep_method(search_dir, search_pattern):
            output_list = []
            DIR_NAME = search_dir
            SEARCH_WORD = search_pattern
            #FLAG_STDOUT = True
            write = sys.stdout.write
            for dirpath, dirs, files in os.walk(DIR_NAME):
                for fn in files:
                    path = os.path.join(dirpath, fn)
                    count = 0
                    enc = 'utf-8'
                    try:
                        for l in codecs.open(path, 'r', enc):
                            count = count + 1
                            if SEARCH_WORD in l:
                                output = ''
                                try:
                                    output = '"' + path + '","' + str(count) + '","' + SEARCH_WORD + '","' + l.replace('"',"'").replace('\r','').replace('\n','') + '"\r\n'
                                    # print('output:', output)
                                    output_list.append(str(path))
                                    output_list.append(str(count))
                                    output_list.append(SEARCH_WORD)
                                    output_list.append(str(l.replace('"',"'").replace('\r','').replace('\n','')))
                                    grep_result.append(output_list)
                                    output_list = []       
                                except:
                                    continue
                                # if FLAG_STDOUT == True:
                                #     write(output)
                                # out.write(output)
                                
                    except:
                        continue

            return grep_result

        #grep_result = grep_method('./', "getTopURLStatus(")

        # ['./testfolder\\testfolder2\\testfile2.txt', '2', 'ge


        # print('This is the start of the start file')

        # print('Fill the from_list list with the py file names that come after from')
        def make_from_list(arg_file):
            if not os.path.exists("./" + arg_file):
                return [[]]
            f = open(arg_file, 'r', encoding='utf-8')
            datalist = f.readlines()

            # Deleted the area surrounded by "" "" ""
            datalist = re.sub(r'\"\"\"(.*?)\"\"\"','',str(datalist))

            # Unification of line breaks
            datalist = datalist.replace('\\r\\n','\\r')
            datalist = datalist.replace('\\n','\\r')

            # Delete everything surrounded by # and \ r
            datalist = re.sub(r'#(.*?)\\r','',datalist)

            while_flag = True
            from_list = []
            from_list_ch = []
            datalist_from_import = datalist
            while(while_flag):
                # Extract the related py file name described in import from from
                # try:
                #     datalist_from_import = datalist_from_import.replace('from ','fffff')
                #     print(datalist_from_import)
                #     pdb.set_trace()
                #     datalist_from_import = datalist_from_import.replace(' import','iiiii')
                # except:
                #     print('from etc. None')
                try:
                    datalist_from_file = (re.search(r'from (.*?) import', datalist_from_import))[0].replace('from ','').replace(' import','')
                    # print('datalist_from_file:', datalist_from_file)
                    # pdb.set_trace()
                    datalist_import_file = (re.search(r' import(.*?)\\r', datalist_from_import))[0].replace('from ','').replace(' import','').replace('\\r','')
                    # print('datalist_import_file:', datalist_import_file)
                    # pdb.set_trace()
                    # print(datalist_import_file)
                    from_list_ch.append(datalist_from_file)
                    from_list_ch.append(datalist_import_file)
                    from_list.append(from_list_ch)
                    from_list_ch = []
                    datalist_from_import = datalist_from_import.replace('from ' + datalist_from_file + ' import', '')
                except:
                    print('-from- py file name -import- no longer exists')
                    while_flag = False
                    break
            f.close()
            return from_list

        from_file_list = make_from_list(start_file)

        return from_file_list

    start_file = arg_1
    from_file_list = from_file_list_export(start_file)

    # (Pdb) from_file_list
    # [['createData', ' getTopURLStatus, getProfileInformation'], ['ccccccc', ' ttttttttt']]
    file_last = open('structure_bone.txt', 'w')
    file_last.write('------------------------------------------------------------' + '\r')
    file_last.write(arg_1 + '\r')


    '''
    Webscrape.py(start file)
        |
        ∟ createDate.py(from information)
        |   |
        |   checkTorihiki(import information)
        |   |
        |   getProfileInformation(import information)
        |
        ∟ ccccccc.py(from information)
            |
            ttttttttt(import information)
    '''
    from_file_list_next = []

    file_last.write('    |' + '\r')
    if from_file_list != []:

        for num, n in enumerate(from_file_list):

            file_last.write('    ∟' + str(from_file_list[int(num)][0]) + '.py\r')
            # Webscrape.py(start file)
            #     |
            #     ∟ createDate.py(from information)
            if ',' in from_file_list[int(num)][1]:

                # When there is more than one import
                import_list = from_file_list[int(num)][1].split(',')

                print('aaa')
                for m in import_list:

                    # After importing into one
                    import_method = m.strip() # Eliminate front and back blanks

                    if int(num) + 1 > len(from_file_list):

                        # When the next from file does not exist
                        file_last.write('        |' + '\r')
                        file_last.write('        ' + str(import_method) + '()\r')

                    else:
                        # When there is the next from file
                        file_last.write('    |    |' + '\r')
                        file_last.write('    |   ' + str(import_method) + '()\r')

                        from_file_list_next = from_file_list_export(str(from_file_list[int(num)][0]) + '.py')
                        # ファイルは存在してる？ ./   core.pdftoxml.py
                        if not os.path.exists('./' + str(from_file_list[int(num)][0]) + '.py'):

                            # no → .がいくつある？  reak.core.pdftoxml
                            if (str(from_file_list[int(num)][0])).count('.') == 0:

                                continue
                            elif (str(from_file_list[int(num)][0])).count('.') == 1:

                                # 1個あるとき　上の階層にその.左側フォルダがある？core
                                if os.path.exists('../../' + (str(from_file_list[int(num)][0])).split('.')[0]):
                                    # 上の階層にあった

                                    print('../../' + (str(from_file_list[int(num)][0])).split('.')[0])
                                    if not os.path.exists('../../' + (str(from_file_list[int(num)][0])).split('.')[0] + '/' +  (str(from_file_list[int(num)][0])).split('.')[1] + '.py'):
                                        continue
                                    else:
                                        with open('../../' + (str(from_file_list[int(num)][0])).split('.')[0] + '/' +  (str(from_file_list[int(num)][0])).split('.')[1] + '.py', encoding='utf-8', errors='ignore') as f:

                                            for line_ch in f:
                                                # print(line_ch)
                                                if from_file_list_next == [[]]:
                                                    continue
                                                for num_ffln, from_file_list_next_ch in enumerate(from_file_list_next):

                                                    if from_file_list_next_ch[1].strip() in line_ch:
                                                        method_name = from_file_list_next_ch[1].strip()
                                                        print('from_file_list_next_ch[1].strip()■■■：', from_file_list_next_ch[1].strip())
                                                        print('line_ch■■■:', line_ch)

                                                        if re.search(rf'\b(?=\w)(.*?){method_name}(.*?)\((.*?)(?!\w)', str(line_ch)):
                                                            # print('Bingo！！！！！！！！！！')

                                                            # In addition. Extracted from commas (because there were things up to parentheses)
                                                            method_text = re.search(rf'\b(?=\w)(.*?){method_name}(.*?)\((.*?)(?!\w)', str(line_ch))

                                                            if int(num_ffln) + 1 == len(from_file_list_next):
                                                                # When the next from file does not exist
                                                                file_last.write('    |        ∟' + str(from_file_list_next_ch[0]) + '.py\r')
                                                                file_last.write('                       |' + '\r')
                                                                file_last.write('                       ' + str(method_name) + '()\r')
                                                            else:
                                                                # When there is the next from file
                                                                file_last.write('    |        ∟' + str(from_file_list_next_ch[0]) + '.py\r')
                                                                file_last.write('    |        |         |' + '\r')
                                                                file_last.write('    |        |         ' + str(method_name) + '()\r')
                                else:
                                    # 上の階層になかった
                                    continue
                            else:
                                # .が2個以上　今の仕様ではスキップ
                                continue
                        # str(from_file_list[int(num)][0]) + '.py' 's str(import_method) + '()' 's Read to the next method
                        print("'./' + str(from_file_list[int(num)][0]) + '.py':", './' + str(from_file_list[int(num)][0]) + '.py')

                        file_name = str(from_file_list[int(num)][0]) + '.py'
                        if file_name[:-3].count('.') == 1:
                            #file_name = '../../' + file_name[:-3].split('.')[0] + '/' + file_name[:-3].split('.')[1] + '.py'
                            path = Path(__file__).parent # swritructure.pyのあるディレクトリ
                            path /= '../../' + file_name[:-3].split('.')[0]
                            file_name = str(path.resolve()) + '\\' + file_name[:-3].split('.')[1] + '.py'

                            #file_name = file_name[:-3].split('.')[1] + '.py'
                            #aaa = sys.path.append(str(Path('__file__').resolve().parent.parent))
                            
                        with open(file_name, encoding='utf-8', errors='ignore') as f:

                            for line_ch in f:
                                # print(line_ch)
                                # If def comes at the beginning, it's coming to the next method, so skip it.
                                if (line_ch.strip())[:3] == 'def':
                                    continue
                                if from_file_list_next == [[]]:
                                    continue
                                for num_ffln, from_file_list_next_ch in enumerate(from_file_list_next):
                                    
                                    if from_file_list_next_ch[1].strip() in line_ch:
                                        method_name = from_file_list_next_ch[1].strip()
                                        # There is a py file name in from in the corresponding line read
                                        #if from_file_list_next_ch[1].strip() + '(' in line_ch:
                                        # When there is something in front, there is a half-width space after the specified method name, and there is something after it
                                        if re.search(rf'\b(?=\w)(.*?){method_name}(.*?)\((.*?)(?!\w)', str(line_ch)):
                                            # print('Bingo！！！！！！！！！！')

                                            # In addition. Extracted from commas (because there were things up to parentheses)
                                            method_text = re.search(rf'\b(?=\w)(.*?){method_name}(.*?)\((.*?)(?!\w)', str(line_ch))

                                            if int(num_ffln) + 1 == len(from_file_list_next):
                                                # When the next from file does not exist
                                                file_last.write('    |    |    |' + '\r')
                                                file_last.write('    |    |    ∟' + str(from_file_list_next_ch[0]) + '.py\r')
                                                file_last.write('    |    |               |' + '\r')
                                                file_last.write('    |    |               ' + str(method_name) + '()\r')
                                            else:
                                                # When there is the next from file
                                                file_last.write('    |    |    |' + '\r')
                                                file_last.write('    |    |    ∟' + str(from_file_list_next_ch[0]) + '.py\r')
                                                file_last.write('    |    |    |          |' + '\r')
                                                file_last.write('    |    |    |          ' + str(method_name) + '()\r')
            else:

                import_method = from_file_list[int(num)][1]

                # After importing into one
                import_method = import_method.strip() # Eliminate front and back blanks

                if int(num) + 1 > len(from_file_list):

                    # When the next from file does not exist
                    file_last.write('        |' + '\r')
                    file_last.write('        ' + str(import_method) + '()\r')

                else:
                    # When there is the next from file
                    file_last.write('    |    |' + '\r')
                    file_last.write('    |   ' + str(import_method) + '()\r')

                    from_file_list_next = from_file_list_export(str(from_file_list[int(num)][0]) + '.py')
                    # ファイルは存在してる？ ./   core.pdftoxml.py
                    print('|||||||||||||||||||||./' + str(from_file_list[int(num)][0]) + '.py')
                    #print('bbbb:', os.path.exists('./' + str(from_file_list[int(num)][0]) + '.py'))

                    if not os.path.exists('./' + str(from_file_list[int(num)][0]) + '.py'):

                        # no → .がいくつある？  reak.core.pdftoxml
                        if (str(from_file_list[int(num)][0])).count('.') == 0:
                            continue
                        elif (str(from_file_list[int(num)][0])).count('.') == 1:

                            # 1個あるとき　上の階層にその.左側フォルダがある？core
                            if os.path.exists('../../' + (str(from_file_list[int(num)][0])).split('.')[0]):
                                # 上の階層にあった

                                print('../../' + (str(from_file_list[int(num)][0])).split('.')[0])
                                if not os.path.exists('../../' + (str(from_file_list[int(num)][0])).split('.')[0] + '/' +  (str(from_file_list[int(num)][0])).split('.')[1] + '.py'):
                                    continue
                                else:
                                    with open('../../' + (str(from_file_list[int(num)][0])).split('.')[0] + '/' +  (str(from_file_list[int(num)][0])).split('.')[1] + '.py', encoding='utf-8', errors='ignore') as f:

                                        for line_ch in f:
                                            # print(line_ch)
                                            if from_file_list_next == [[]]:
                                                continue
                                            for num_ffln, from_file_list_next_ch in enumerate(from_file_list_next):

                                                if from_file_list_next_ch[1].strip() in line_ch:

                                                    method_name = from_file_list_next_ch[1].strip()
                                                    print('from_file_list_next_ch[1].strip()■■■：', from_file_list_next_ch[1].strip())
                                                    print('line_ch■■■:', line_ch)

                                                    if re.search(rf'\b(?=\w)(.*?){method_name}(.*?)\((.*?)(?!\w)', str(line_ch)):
                                                        # print('Bingo！！！！！！！！！！')

                                                        # In addition. Extracted from commas (because there were things up to parentheses)
                                                        method_text = re.search(rf'\b(?=\w)(.*?){method_name}(.*?)\((.*?)(?!\w)', str(line_ch))

                                                        if int(num_ffln) + 1 == len(from_file_list_next):
                                                            # When the next from file does not exist
                                                            file_last.write('    |        ∟' + str(from_file_list_next_ch[0]) + '.py\r')
                                                            file_last.write('                       |' + '\r')
                                                            file_last.write('                       ' + str(method_name) + '()\r')
                                                        else:
                                                            # When there is the next from file
                                                            file_last.write('    |        ∟' + str(from_file_list_next_ch[0]) + '.py\r')
                                                            file_last.write('    |        |         |' + '\r')
                                                            file_last.write('    |        |         ' + str(method_name) + '()\r')
                            else:
                                # 上の階層になかった
                                continue
                        else:
                            # .が2個以上　今の仕様ではスキップ
                            continue

                    else:
                        with open('./' + str(from_file_list[int(num)][0]) + '.py', encoding='utf-8', errors='ignore') as f:
                            
                            for line_ch in f:
                                # print(line_ch)

                                for num_ffln, from_file_list_next_ch in enumerate(from_file_list_next):

                                    if from_file_list_next_ch[1].strip() in line_ch:
                                        method_name = from_file_list_next_ch[1].strip()
                                        print('from_file_list_next_ch[1].strip()■■■：', from_file_list_next_ch[1].strip())
                                        print('line_ch■■■:', line_ch)

                                        if re.search(rf'\b(?=\w)(.*?){method_name}(.*?)\((.*?)(?!\w)', str(line_ch)):
                                            # print('Bingo！！！！！！！！！！')

                                            # In addition. Extracted from commas (because there were things up to parentheses)
                                            method_text = re.search(rf'\b(?=\w)(.*?){method_name}(.*?)\((.*?)(?!\w)', str(line_ch))

                                            if int(num_ffln) + 1 == len(from_file_list_next):
                                                # When the next from file does not exist
                                                file_last.write('    |        ∟' + str(from_file_list_next_ch[0]) + '.py\r')
                                                file_last.write('                       |' + '\r')
                                                file_last.write('                       ' + str(method_name) + '()\r')
                                            else:
                                                # When there is the next from file
                                                file_last.write('    |        ∟' + str(from_file_list_next_ch[0]) + '.py\r')
                                                file_last.write('    |        |         |' + '\r')
                                                file_last.write('    |        |         ' + str(method_name) + '()\r')

                # When there is only one import
                import_method = from_file_list[int(num)][1].strip() # Eliminate front and back blanks
    
                if int(num) + 1 == len(from_file_list):
                    # When the next from file does not exist
                    file_last.write('        |' + '\r')
                    file_last.write('        ' + str(import_method) + '()\r')
                else:
                    # When there is the next from file
                    file_last.write('    |   |' + '\r')
                    file_last.write('    |   ' + str(import_method) + '()\r')

    file_last.close()
    print('◆◆◆◆◆◆structure_bone.txt is output. Please confirm.◆◆◆◆◆◆')

    link_list = []
    with open('./structure_bone.txt', newline='') as test_file:
        str_data_list = list(csv.reader(test_file, delimiter='~'))

    # '∟'がいくつあるかで状況判断　1個だけそのファイルの横最大文字列測る（ファイル名探、メソッド名探）　2個だけのときも同様　3個　4個　まず全貼り付け範囲確認
    #　一個の時のを縦の順で貼り付け、次に二個のを順で貼り付け　3個　4個
    #　メソッド名で線で紐づけ描写
    second_file_name = []
    first_method_name = []
    third_file_name = []
    second_method_name = []

    # structure_bone.txtからファイル名とメソッド名の位置情報と何番目に出現かを抽出
    for dl_num, record in enumerate(str_data_list):
        if dl_num == 0 or dl_num == 1:
            continue
        if (record[0].replace('    ∟','')).count(' ') == 0:
            second_file_name.append(record[0].replace('    ∟',''))
        if (record[0].replace('    |   ','')).count(' ') == 0:
            first_method_name.append(record[0].replace('    |   ','').replace('()',''))
        if (record[0].replace('    |    |    ∟','')).count(' ') == 0:
            third_file_name.append(record[0].replace('    |    |    ∟',''))
        if (record[0].replace('    |    |    |          ','')).count(' ') == 0:
            if not '()' in record[0].replace('    |    |    |          ',''):
                continue
            second_method_name.append(record[0].replace('    |    |    |          ','').replace('()',''))

    all_file_name_ch = []
    all_file_name = []
    # 上記抽出したものを2次元リストへ
    for sfn_num, second_file_name_in in enumerate(second_file_name):
        print('second file:', second_file_name_in + ' ' + str(sfn_num + 1))
        all_file_name_ch.append('second file:')
        all_file_name_ch.append(second_file_name_in)
        all_file_name_ch.append(str(sfn_num + 1))
        all_file_name.append(all_file_name_ch)
        all_file_name_ch = []
    for sfn_num, first_method_name_in in enumerate(first_method_name):
        print('first method:', first_method_name_in + ' ' + str(sfn_num + 1))
        all_file_name_ch.append('first method:')
        all_file_name_ch.append(first_method_name_in)
        all_file_name_ch.append(str(sfn_num + 1))
        all_file_name.append(all_file_name_ch)
        all_file_name_ch = []
    for sfn_num, third_file_name_in in enumerate(third_file_name):
        print('third file:', third_file_name_in + ' ' + str(sfn_num + 1))
        all_file_name_ch.append('third file:')
        all_file_name_ch.append(third_file_name_in)
        all_file_name_ch.append(str(sfn_num + 1))
        all_file_name.append(all_file_name_ch)
        all_file_name_ch = []
    for sfn_num, second_method_name_in in enumerate(second_method_name):
        print('second method:', second_method_name_in + ' ' + str(sfn_num + 1))
        all_file_name_ch.append('second method:')
        all_file_name_ch.append(second_method_name_in)
        all_file_name_ch.append(str(sfn_num + 1))
        all_file_name.append(all_file_name_ch)
        all_file_name_ch = []



    def file_paste(link_list, wb, data_list, file_name, start_r, max_char_len, start_c, file_kind):
        ori_data_list = data_list
        # シートを取得
        ws = wb['Sheet']
        # legth research
        data_list = []
        try:
            print('file_name:', file_name)



            if file_name[:-3].count('.') == 1:

                #file_name = '../../' + file_name[:-3].split('.')[0] + '/' + file_name[:-3].split('.')[1] + '.py'
                path = Path(__file__).parent # swritructure.pyのあるディレクトリ
                path /= '../../' + file_name[:-3].split('.')[0]
                file_name = str(path.resolve()) + '\\' + file_name[:-3].split('.')[1] + '.py'
                #file_name = file_name[:-3].split('.')[1] + '.py'

                #aaa = sys.path.append(str(Path('__file__').resolve().parent.parent))

            with open(file_name, newline='', encoding='utf-8') as test_file:
                data_list = list(csv.reader(test_file, delimiter='~'))
                print('data_list[0]:', data_list[0])

        except:
            print('third file open error')
            return wb, ori_data_list, start_r, max_char_len, start_c

        # list []削除
        new_data_list = []
        for data_list_in in data_list:
            if data_list_in != []:
                new_data_list.append(data_list_in)
        data_list = new_data_list
        
        data_list.insert(0,[file_name])

        # データの書き込み
        for record in data_list:
            if max_char_len < int(len(str(record)) / 10) + 3:
                max_char_len = int(len(str(record)) / 10) + 3

        # horizontalは調整いる
        # max_char_len = int(max_char_len / 10) - 1
        print('max_char_len:', max_char_len)
        # print(str(max_char_len)) # horizontal
        # print(str(len(data_list))) # vertical
        
        max_rows_len = start_r + int(len(data_list))

        # 指定背景色 https://pg-chain.com/python-excel-color
        new_fill = px.styles.PatternFill(patternType='solid', fgColor='000000', bgColor='000000')
        # 指定文字色
        new_font = px.styles.fonts.Font(color='FFFFFF')


        for dl_num, data_list_in in enumerate(data_list):
            # if data_list_in == []:
            #     ws.cell(row=start_r+3+dl_num,column=start_c).value = ''
            #     ws.cell(row=start_r+3+dl_num,column=start_c).font = new_font
            #     for max_char_len_in in range(max_char_len):
            #         ws.cell(row=start_r+3+dl_num, column=start_c+max_char_len_in).fill = new_fill
            #     continue
            if dl_num == 0:
                ws.cell(row=start_r+3+dl_num,column=start_c).value = data_list_in[0]
                continue        
            # print('data_list_in[0]:', data_list_in[0])
            # pdb.set_trace()
            ws.cell(row=start_r+3+dl_num,column=start_c).value = data_list_in[0]
            # above file under method  ['second method:', 'checkGaiyouGessu', '35']
            for all_file_name_in in all_file_name:
                if file_kind == 1:
                    # first(main) fileのとき
                    if 'first method:'== all_file_name_in[0]:
                        method_arg = all_file_name_in[1] + '('
                        for method_arg in data_list_in[0]:
                            link_list.append('1')
                            link_list.append(all_file_name_in[1])

                            link_list.append(str(dl_num))
                elif file_kind == 2:
                    # second fileのとき

                    if 'second method:'== all_file_name_in[0]:
                        method_arg = all_file_name_in[1] + '('
                        for method_arg in data_list_in[0]:
                            link_list.append('2')
                            link_list.append(all_file_name_in[1])
                            link_list.append(str(dl_num))
                elif file_kind == 3:
                    # second fileのとき
                    if 'third method:'== all_file_name_in[0]:
                        method_arg = all_file_name_in[1] + '('
                        for method_arg in data_list_in[0]:
                            link_list.append('3')
                            link_list.append(all_file_name_in[1])
                            link_list.append(str(dl_num))

            ws.cell(row=start_r+3+dl_num,column=start_c).font = new_font
            for max_char_len_in in range(max_char_len):
                ws.cell(row=start_r+3+dl_num, column=start_c+max_char_len_in).fill = new_fill
            
        wb.save('./structure.xlsx')

        print('これがサードファイルの右列', max_char_len)
        print('これがサードファイル貼り付け済みの行数:', max_rows_len)
        print('ｔｔｔｔｔｔｔ')
        return wb, data_list, max_rows_len, max_char_len, start_c

    data_list = []
    # ブックを取得
    wb = px.Workbook()
    file_name = arg_1 # 初めだけ指定　arg_1

    # wb, data_list, max_char_len = main_file_paste(wb, data_list, file_name, 1, 1)
    start_c = 1
    max_rows_len = 1
    max_char_len = 0
    file_kind = 1 # first(main) file → 1, second file → 2, third file → 3, fourth file → 4
    wb, data_list, max_rows_len, max_char_len, start_c = file_paste(link_list, wb, data_list, file_name, max_rows_len, max_char_len, start_c, file_kind)

    main_last_char_len = max_char_len
    wb.close()

    wb = px.load_workbook('./structure.xlsx')

    # from second file
    max_rows_len = 1
    flag_first_second = 0
    flag_first_third = 0

    for all_file_name_in in all_file_name:
        #first_method_count = 0
        if all_file_name_in[0] == 'second file:':
            # 1回目だけスタートの始点（横）の調整を行う

            if flag_first_second == 0:
                flag_first_second += 1
                start_c = main_last_char_len + 1
                max_rows_len = 1
                max_char_len = 0
            file_kind = 2
            print('max_rows_len', max_rows_len)
            print('max_char_len', max_char_len)
            # second file の列のコードを追加していく
            wb, data_list, max_rows_len, max_char_len, start_c = file_paste(link_list, wb, data_list, all_file_name_in[1], max_rows_len, max_char_len, start_c, file_kind)

        if all_file_name_in[0] == 'third file:':
            # 1回目だけスタートの始点（横）の調整を行う
            if flag_first_third == 0:
                flag_first_third += 1
                second_last_char_len = max_char_len
                start_c = main_last_char_len + second_last_char_len + 1
                max_rows_len = 1
                max_char_len = 0
            file_kind = 3
            print('max_rows_len', max_rows_len)
            print('max_char_len', max_char_len)
            # second file の列のコードを追加していく
            wb, data_list, max_rows_len, max_char_len, start_c = file_paste(link_list, wb, data_list, all_file_name_in[1], max_rows_len, max_char_len, start_c, file_kind)

    # side = Side(style='thin')
    # border = Border(top=side, bottom=side, left=side, right=side)
    # font_title = Font(name='メイリオ', size=20)


    # ws['A1'].font = font_title

    # ws['A2'] = '<NEW>'
    # ws['A2'].font = font_title

    # ws.cell(row=len(df_new)+5, column=1).value = '<WON>'
    # ws.cell(row=len(df_new)+5, column=1).font = font_title
    # for i in range(1, len(ws.max_column)+1): # ここがうまくいきません
    #     ws.cell(row=len(df_new)+6, columnns=i).border = border

if __name__ == "__main__":
    args = sys.argv
    # 第一引数
    # try:
    arg_1 = args[1]
    get_structure(arg_1)
    # except:
    #     print('Please write "python swritructure <your main filename>" or other reason')
